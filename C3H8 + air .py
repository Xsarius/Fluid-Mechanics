import cantera as ct
from openpyxl import load_workbook
from sdtoolbox.postshock import CJspeed

P1 = ct.one_atm
T1 = 300
U = 2000
mech = 'GRI30.yaml'
M_air = 0.01778
rho_air = 1.2

b = 5   # coeff dependend on fuel ; b = <num of C> * 1 + <num of H> * 1/4 - <num of O> * 1/2
M_fuel = 0.7215 
rho_fuel = 626
flammabillityRange = [0.021, 0.101] # volumetric flammabillity range

molarFlammabillityRange = [round((rho_fuel*flammabillityRange[0]/M_fuel)/(rho_fuel*flammabillityRange[0]/M_fuel+rho_air*(1 - flammabillityRange[0])/M_air), 2),
                                round((rho_fuel*flammabillityRange[1]/M_fuel)/(rho_fuel*flammabillityRange[1]/M_fuel+rho_air*(1 - flammabillityRange[1])/M_air), 2)]

print(molarFlammabillityRange)

wb = load_workbook("Data.xlsx")
ws = wb.create_sheet("C3H8 + air")

ws["A1"] = "CJ speed"
ws["B1"] = "fuel molar frac"
ws["C1"] = "equvalence ratio"
ws["D1"] = "Flammabillity range low"
ws["E1"] = "Flammabillity range high"
i = 1 

for p in range(1,100):

    equivalenceRatio = p/10
    fuelMoleFrac = equivalenceRatio/(b+equivalenceRatio)

    if round(fuelMoleFrac,2) > molarFlammabillityRange[0] and round(fuelMoleFrac,2) < molarFlammabillityRange[1]:
        q = 'C3H8:'+ str(equivalenceRatio) +' O2:'+ str(b) +' N2:' + str(b * 3.76)
        gas = ct.Solution(mech)

        ws["A" + str(i + 1)] = round(CJspeed(P1, T1, q, mech),2)
        ws["B" + str(i + 1)] = round(fuelMoleFrac,2)
        ws["C" + str(i + 1)] = equivalenceRatio
        ws["D" + str(i + 1)] = molarFlammabillityRange[0]
        ws["E" + str(i + 1)] = molarFlammabillityRange[1]
        i = i + 1

    elif round(fuelMoleFrac,2) > molarFlammabillityRange[1]:
        break

wb.save("Data.xlsx")
    