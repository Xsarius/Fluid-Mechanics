import cantera as ct
from openpyxl import load_workbook
from sdtoolbox.postshock import CJspeed

P1 = ct.one_atm
T1 = 300
U = 2000
mech = 'GRI30.yaml'
M_air = 0.032
rho_air = 1.429

b = 0.5   # coeff dependend on fuel ; b = <num of C> * 1 + <num of H> * 1/4 - <num of O> * 1/2
M_fuel = 0.001 
rho_fuel = 0.0899
flammabillityRange = [0.038,0.95] # molar flammabillity range

molarFlammabillityRange = flammabillityRange

print(molarFlammabillityRange)

wb = load_workbook("Data.xlsx")
ws = wb.create_sheet("H2 + O2")

ws["A1"] = "CJ speed"
ws["B1"] = "fuel molar frac"
ws["C1"] = "equvalence ratio"
ws["D1"] = "Flammabillity range low"
ws["E1"] = "Flammabillity range high"
i = 1
    
for p in range(1,21):

    equivalenceRatio = p/10
    fuelMoleFrac = equivalenceRatio/(b+equivalenceRatio)

    if round(fuelMoleFrac,2) > molarFlammabillityRange[0] and round(fuelMoleFrac,2) < molarFlammabillityRange[1]:
        q = 'H2:'+ str(equivalenceRatio) +' O2:'+ str(b)
        gas = ct.Solution(mech)

        ws["A" + str(i + 1)] = round(CJspeed(P1, T1, q, mech),2)
        ws["B" + str(i + 1)] = round(fuelMoleFrac,2)
        ws["C" + str(i + 1)] = equivalenceRatio
        ws["D" + str(i + 1)] = molarFlammabillityRange[0]
        ws["E" + str(i + 1)] = molarFlammabillityRange[1]
        i = i + 1

    elif round(fuelMoleFrac,2) > molarFlammabillityRange[1]:
        break

wb.save("Data.xlsx")
    